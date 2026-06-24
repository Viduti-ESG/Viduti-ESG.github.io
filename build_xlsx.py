import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

POSTS = json.load(open(r"c:\Viduti\esg-site\posts_extracted.json", encoding="utf-8"))
DATA = {p["slug"]: p for p in POSTS}
CLUST = json.load(open(r"c:\Viduti\esg-site\clusters.json"))

def rep(members):
    return max(members, key=lambda s: sum(len(x["t"]) for x in DATA[s]["sections"]))

raw = open(r"c:\Viduti\esg-site\linkedin_posts_content.txt", encoding="utf-8").read()
topics = []
for block in raw.split("===TOPIC===")[1:]:
    block = block.split("===FILE END===")[0]
    tid = re.search(r"id:\s*(.+)", block).group(1).strip()
    posts = []
    for pb in block.split("===POST===")[1:]:
        pb = pb.split("===ENDTOPIC===")[0]
        body = pb.split("===ENDPOST===")[0]
        status = re.search(r"status:\s*(.+)", body).group(1).strip()
        notes = re.search(r"notes:\s*(.+)", body).group(1).strip()
        text = body.split("text:\n", 1)[1].strip()
        posts.append({"status": status, "notes": notes, "text": text})
    topics.append({"id": tid, "posts": posts})

order = list(CLUST.keys())
topics.sort(key=lambda t: order.index(t["id"]))

wb = Workbook()
ws = wb.active
ws.title = "LinkedIn Posts"

GREEN = "10b981"; DARK = "0f172a"; LIGHT = "ECFDF5"; AMBER = "FEF3C7"
hdr_fill = PatternFill("solid", fgColor=DARK)
hdr_font = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)

cols = ["#", "Topic / Theme", "Category", "Source posts", "Live URL (representative)",
        "Post", "LinkedIn post (ready to paste)", "Char count", "Fact-check status",
        "Fact-check notes & corrections"]
ws.append(cols)
for c in range(1, len(cols) + 1):
    cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border

def status_fill(s):
    sl = s.lower()
    if "major" in sl or "low news" in sl or "flag" in sl:
        return PatternFill("solid", fgColor=AMBER)
    if "correction" in sl:
        return PatternFill("solid", fgColor=LIGHT)
    return PatternFill("solid", fgColor="FFFFFF")

r = 2; tnum = 0
for t in topics:
    tnum += 1
    members = CLUST[t["id"]]
    repslug = rep(members)
    meta = DATA[repslug]
    title = meta["title"]; cat = meta["category"]
    url = f"https://greencurve.solutions/posts/{repslug}"
    nmem = len(members); npost = len(t["posts"]); start = r
    for i, p in enumerate(t["posts"], 1):
        ws.cell(r, 1, tnum); ws.cell(r, 2, title); ws.cell(r, 3, cat)
        ws.cell(r, 4, str(nmem)); ws.cell(r, 5, url)
        ws.cell(r, 6, f"{i} of {npost}" if npost > 1 else "1")
        ws.cell(r, 7, p["text"]); ws.cell(r, 8, len(p["text"]))
        sc = ws.cell(r, 9, p["status"]); sc.fill = status_fill(p["status"])
        ws.cell(r, 10, p["notes"])
        for c in range(1, 11):
            cl = ws.cell(r, c); cl.border = border
            cl.alignment = center if c in (1, 4, 6, 8) else wrap_top
        r += 1
    if npost > 1:
        for c in (1, 2, 3, 4, 5):
            ws.merge_cells(start_row=start, start_column=c, end_row=r - 1, end_column=c)

widths = {1: 5, 2: 34, 3: 18, 4: 8, 5: 40, 6: 7, 7: 95, 8: 8, 9: 22, 10: 52}
for k, v in widths.items():
    ws.column_dimensions[get_column_letter(k)].width = v
ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"
for rr in range(2, r):
    txt = ws.cell(rr, 7).value or ""
    lines = max(len(txt) // 90 + txt.count("\n") + 2, 6)
    ws.row_dimensions[rr].height = min(lines * 14, 620)

ws2 = wb.create_sheet("README & Methodology", 0)
ws2.column_dimensions['A'].width = 115
lines = [
 ("Green Curve - LinkedIn Posts for the Insights / News Blog", True, 15, GREEN),
 ("Analytical, non-promotional LinkedIn posts written from the Green Curve Insights articles. Generated 2026-06-22.", False, 11, "475569"),
 ("", False, 11, "000000"),
 ("WHAT THIS IS", True, 12, "0f172a"),
 ("- 93 published blog posts were de-duplicated into 53 unique topics (several articles were near-identical regenerated variants).", False, 11, "000000"),
 ("- Each topic has one or more LinkedIn post, written to be analytical and informative - NOT marketing. There are no 'try our tool' calls-to-action; the posts stand on insight alone.", False, 11, "000000"),
 ("- Every factual claim was checked. Where the source article was outdated or wrong, the post was corrected and the change is logged in the 'Fact-check notes' column.", False, 11, "000000"),
 ("", False, 11, "000000"),
 ("HOW TO USE", True, 12, "0f172a"),
 ("- The 'LinkedIn Posts' tab is the working sheet: copy the 'LinkedIn post' column straight into LinkedIn.", False, 11, "000000"),
 ("- Char count is shown so you can gauge length (LinkedIn truncates with 'see more' around 1,300 characters; all posts deliver value above the fold).", False, 11, "000000"),
 ("- 'Fact-check status' colour key: white = verified as-is; light green = verified WITH corrections applied; amber = contains a MAJOR correction, a flag, or low underlying news value (read the notes before posting).", False, 11, "000000"),
 ("", False, 11, "000000"),
 ("KEY FACT CORRECTIONS APPLIED (web-verified, June 2026)", True, 12, "B45309"),
 ("1. IMO Net-Zero Framework: source said 'postponed to April 2026'. CORRECTED - adoption was ADJOURNED FOR ONE YEAR at the Oct 2025 extraordinary session (to OCTOBER 2026), amid US opposition.", False, 11, "000000"),
 ("2. EU CSRD: sources described the ORIGINAL scope. CORRECTED - the Omnibus package (stop-the-clock Apr 2025; substance adopted Feb 2026) raised the threshold to 1,000+ employees AND EUR 450m+ turnover and exempted listed SMEs.", False, 11, "000000"),
 ("3. EU Deforestation Regulation (EUDR): source said enforcement 'Dec 2025'. CORRECTED - main obligations for large operators now apply from 30 December 2026 (30 June 2027 for micro/small).", False, 11, "000000"),
 ("4. India non-fossil power capacity: source said '40% crossed in 2023'. CORRECTED - 40% reached in 2021; 50% crossed in mid-2025, ~5 years early.", False, 11, "000000"),
 ("5. Climate Action 100+: source said '$65 trillion'. CORRECTED - over $68 trillion AUM across 700+ investors, ~170 focus companies.", False, 11, "000000"),
 ("6. COP host: source referenced 'COP31 Brazil'. CORRECTED - COP30 was Belem, Brazil (Nov 2025); COP31 is hosted by Turkey/Antalya (Nov 2026), Australia leading negotiations.", False, 11, "000000"),
 ("7. Fossil-fuel EXPORT emissions: a source implied India is a top fossil-fuel exporter. CLARIFIED - India is a major importer/consumer & domestic coal producer, not a leading exporter; post reframed.", False, 11, "000000"),
 ("8. EPR market sizes, '6 lakh fake certificates', '3.7 mn tonnes traded', '11 gold recyclers', India e-waste tonnage: flagged as industry estimates / directional and hedged in the posts.", False, 11, "000000"),
 ("", False, 11, "000000"),
 ("NOTE", True, 12, "0f172a"),
 ("Stable, well-established facts (rule notification dates, GHG Protocol/GRI/ISSB/SBTi/TNFD fundamentals, SEBI circular dates) were verified against the official notifications cited in each article and public record, and are marked 'Verified'.", False, 11, "475569"),
 ("The 'MoEFCC internal appointment' topic has thin underlying news value and is flagged as optional/skip for publication.", False, 11, "475569"),
]
for i, (txt, b, sz, col) in enumerate(lines, 1):
    cell = ws2.cell(i, 1, txt); cell.font = Font(bold=b, size=sz, color=col)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
ws2.sheet_view.showGridLines = False

out = r"c:\Viduti\esg-site\Green_Curve_LinkedIn_Posts.xlsx"
wb.save(out)
total_posts = sum(len(t["posts"]) for t in topics)
print("Saved:", out)
print("Topics:", len(topics), " Posts:", total_posts)
for t in topics:
    for i, p in enumerate(t["posts"], 1):
        if len(p["text"]) < 300:
            print("SHORT:", t["id"], i, len(p["text"]))
