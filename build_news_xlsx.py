import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TOPICS = json.load(open(r"c:\Viduti\esg-site\news_topics.json", encoding="utf-8"))
BYKEY = {t["key"]: t for t in TOPICS}

def parse(path):
    raw = open(path, encoding="utf-8").read()
    d = {}
    for block in raw.split("===TOPIC===")[1:]:
        block = block.split("===FILE END===")[0]
        tid = re.search(r"id:\s*(.+)", block).group(1).strip()
        pb = block.split("===POST===")[1].split("===ENDPOST===")[0]
        status = re.search(r"status:\s*(.+)", pb).group(1).strip()
        notes = re.search(r"notes:\s*(.+)", pb).group(1).strip()
        text = pb.split("text:\n", 1)[1].strip()
        d[tid] = {"status": status, "notes": notes, "text": text}
    return d

NEW = parse(r"c:\Viduti\esg-site\news_posts_content.txt")
OLD = parse(r"c:\Viduti\esg-site\linkedin_posts_content.txt")

# reused key -> old id
REUSE = {
 "n10": "digest_20may", "n16": "digest_22may", "n21": "digest_26may",
 "n46": "ndc_gap", "n47": "fossil_export_emissions", "n48": "imo_mepc83",
 "n49": "cat_power_sector", "n50": "fossil_import_trap", "n51": "imo_netzero_postpone",
 "n52": "ct_ca100", "n53": "ct_gas_network", "n54": "uniper", "n55": "uk_transition_plans",
 "n56": "sasb_oilgas", "n57": "ct_recalibrating", "n58": "cat_bonn",
}

def get_post(key):
    if key in NEW:
        return NEW[key]
    return OLD[REUSE[key]]

wb = Workbook()
ws = wb.active
ws.title = "News LinkedIn Posts"
GREEN="10b981"; DARK="0f172a"; LIGHT="ECFDF5"; AMBER="FEF3C7"
hdr_fill=PatternFill("solid",fgColor=DARK); hdr_font=Font(color="FFFFFF",bold=True,size=11)
thin=Side(style="thin",color="D1D5DB"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap_top=Alignment(wrap_text=True,vertical="top"); center=Alignment(horizontal="center",vertical="top",wrap_text=True)

cols=["#","Date","News headline (blog title)","# source\nposts","Live URL","LinkedIn post (ready to paste)","Char\ncount","Fact-check status","Fact-check notes & corrections"]
ws.append(cols)
for c in range(1,len(cols)+1):
    cell=ws.cell(1,c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=border

def status_fill(s):
    sl=s.lower()
    if "major" in sl or "low news" in sl: return PatternFill("solid",fgColor=AMBER)
    if "correction" in sl or "flag" in sl: return PatternFill("solid",fgColor=LIGHT)
    return PatternFill("solid",fgColor="FFFFFF")

r=2
for i in range(1,59):
    key=f"n{i:02d}"
    t=BYKEY[key]; p=get_post(key)
    url=f"https://greencurve.solutions/posts/{t['rep']}"
    ws.cell(r,1,i); ws.cell(r,2,t["date"]); ws.cell(r,3,t["title"])
    ws.cell(r,4,str(len(t["members"]))); ws.cell(r,5,url)
    ws.cell(r,6,p["text"]); ws.cell(r,7,len(p["text"]))
    sc=ws.cell(r,8,p["status"]); sc.fill=status_fill(p["status"])
    ws.cell(r,9,p["notes"])
    for c in range(1,10):
        cl=ws.cell(r,c); cl.border=border
        cl.alignment=center if c in (1,2,4,7) else wrap_top
    r+=1

widths={1:5,2:12,3:40,4:8,5:40,6:95,7:8,8:24,9:52}
for k,v in widths.items(): ws.column_dimensions[get_column_letter(k)].width=v
ws.row_dimensions[1].height=32; ws.freeze_panes="A2"
for rr in range(2,r):
    txt=ws.cell(rr,6).value or ""
    ws.row_dimensions[rr].height=min(max(len(txt)//90+txt.count("\n")+2,6)*14,640)

# README
ws2=wb.create_sheet("README & Methodology",0)
ws2.column_dimensions['A'].width=118
lines=[
 ("Green Curve - LinkedIn Posts for the NEWS blog (Daily Digest / 'Latest Insights')",True,15,GREEN),
 ("Analytical, non-promotional LinkedIn posts written from the Green Curve NEWS feed - the 'Daily Digest' category ('in-depth regulatory analysis, published every morning'). Generated 2026-06-23.",False,11,"475569"),
 ("",False,11,"000000"),
 ("SCOPE",True,12,"0f172a"),
 ("- Source: the live posts index (Viduti-ESG GitHub repo) holds 147 published posts; exactly 63 are category 'Daily Digest' (the news feed). These 63 were de-duplicated into 58 unique news topics.",False,11,"000000"),
 ("- This is the NEWS deliverable only. The compliance/explainer blogs (E-Waste/Battery/Plastic rules, framework guides, etc.) are covered in the separate Green_Curve_LinkedIn_Posts.xlsx.",False,11,"000000"),
 ("- 3 topics bundled near-identical variants: SBTi strategy reset (3 posts), Carbon Tracker gas-network trilogy Snam/Italgas/Enagas (3), CA100+ audit+methodology (2). '# source posts' shows how many underlying posts each topic covers.",False,11,"000000"),
 ("",False,11,"000000"),
 ("HOW TO USE",True,12,"0f172a"),
 ("- Copy the 'LinkedIn post' column straight into LinkedIn. Posts are written for the blog's own audience: Indian CFOs, CSOs, boards and ESG investors - analysis and India angle, no marketing CTAs.",False,11,"000000"),
 ("- Fact-check status colours: white = verified as-is; light green = verified with corrections/flags applied; amber = major correction or low news value (read the note first).",False,11,"000000"),
 ("- 'Verified (as reported)' = corporate deal/fund/M&A items whose specifics (amounts, names, dates) come from the source news article (mostly esgtoday.com / GreenBiz-Trellis). The post attributes them and the analytical India angle is original. Policy items were independently web-verified.",False,11,"000000"),
 ("",False,11,"000000"),
 ("KEY FACT CORRECTIONS / VERIFICATIONS (web-checked, June 2026)",True,12,"B45309"),
 ("- IMO Net-Zero Framework (n51): source title says 'Postponement to April 2026'. CORRECTED in-post - adoption was ADJOURNED ONE YEAR at the Oct 2025 extraordinary session, to OCTOBER 2026 (amid US opposition).",False,11,"000000"),
 ("- EUDR (n27, n41): source repeated 'Dec 2024/2025' enforcement. CORRECTED - large-operator obligations now apply 30 Dec 2026 (30 Jun 2027 micro/small).",False,11,"000000"),
 ("- California SB 253 (n31): VERIFIED - Scope 1&2 first reporting deadline 10 Aug 2026, Scope 3 from 2027, >$1bn global revenue (as amended by SB 219).",False,11,"000000"),
 ("- EU ECGT anti-greenwashing (n42): VERIFIED - 20-state infringement (late May 2026); transposition due 27 Mar 2026; applies 27 Sep 2026; bans offset-based 'carbon neutral' claims.",False,11,"000000"),
 ("- UK 7th Carbon Budget (n44): VERIFIED - 535 MtCO2e (2038-42), 87% cut by 2040 vs 1990, must be legislated by 30 Jun 2026.",False,11,"000000"),
 ("- New York CLCPA (n34): VERIFIED - 40%-by-2030 replaced with 60%-by-2040; regs by 2028; GWP accounting shifted 20yr->100yr (signed May 2026).",False,11,"000000"),
 ("- India non-fossil capacity (n03 and reused NDC posts): UPDATED - India crossed 50% non-fossil installed capacity in mid-2025, ahead of schedule.",False,11,"000000"),
 ("- Climate Action 100+ (n52): figure corrected to $68 trillion AUM / ~170 focus companies.",False,11,"000000"),
 ("",False,11,"000000"),
 ("REUSED POSTS",True,12,"0f172a"),
 ("16 topics (n10, n16, n21, n46-n58) overlap with the earlier deliverable's news items (the 5 June and 12 June Carbon Tracker/CAT/IMO/NDC pieces and the three EPR digests). Their posts are carried over verbatim, already fact-checked.",False,11,"475569"),
]
for i,(txt,b,sz,col) in enumerate(lines,1):
    cell=ws2.cell(i,1,txt); cell.font=Font(bold=b,size=sz,color=col); cell.alignment=Alignment(wrap_text=True,vertical="top")
ws2.sheet_view.showGridLines=False

out=r"c:\Viduti\LinkedIn Posts\Green_Curve_News_LinkedIn_Posts.xlsx"
wb.save(out)
print("Saved:",out)
print("Rows:",r-2)
short=[(f"n{i:02d}",len(get_post(f'n{i:02d}')['text'])) for i in range(1,59) if len(get_post(f'n{i:02d}')['text'])<300]
print("short posts:",short)
